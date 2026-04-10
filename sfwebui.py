# -*- coding: utf-8 -*-
# -----------------------------------------------------------------
# Name:         sfwebui
# Purpose:      User interface class for use with a web browser
#
# Author:       Steve Micallef <steve@binarypool.com>
#
# Created:      30/09/2012
# Copyright:    (c) Steve Micallef 2012
# License:      MIT
# -----------------------------------------------------------------
import csv
import html
import ipaddress
import json
import logging
import multiprocessing as mp
import random
import re
import string
import time
from copy import deepcopy
from io import BytesIO, StringIO
from operator import itemgetter
from pathlib import Path

import cherrypy
from cherrypy import _cperror

from mako.lookup import TemplateLookup
from mako.template import Template

import openpyxl

import secure

from sflib import SpiderFoot

from sfscan import startSpiderFootScanner

from spiderfoot import SpiderFootDb
from spiderfoot import SpiderFootHelpers
from spiderfoot import __version__
from spiderfoot.geolite import GeoLiteWorkspace
from spiderfoot.logger import logListenerSetup, logWorkerSetup
from spiderfoot.validators import FindingValidatorEngine

mp.set_start_method("spawn", force=True)


class SpiderFootWebUi:
    """SpiderFoot web interface."""

    lookup = TemplateLookup(directories=[''])
    defaultConfig = dict()
    config = dict()
    token = None
    docroot = ''

    def __init__(self: 'SpiderFootWebUi', web_config: dict, config: dict, loggingQueue: 'logging.handlers.QueueListener' = None) -> None:
        """Initialize web server.

        Args:
            web_config (dict): config settings for web interface (interface, port, root path)
            config (dict): SpiderFoot config
            loggingQueue: TBD

        Raises:
            TypeError: arg type is invalid
            ValueError: arg value is invalid
        """
        if not isinstance(config, dict):
            raise TypeError(f"config is {type(config)}; expected dict()")
        if not config:
            raise ValueError("config is empty")

        if not isinstance(web_config, dict):
            raise TypeError(f"web_config is {type(web_config)}; expected dict()")
        if not config:
            raise ValueError("web_config is empty")

        self.docroot = web_config.get('root', '/').rstrip('/')

        # 'config' supplied will be the defaults, let's supplement them
        # now with any configuration which may have previously been saved.
        self.defaultConfig = deepcopy(config)
        dbh = SpiderFootDb(self.defaultConfig, init=True)
        sf = SpiderFoot(self.defaultConfig)
        self.config = sf.configUnserialize(dbh.configGet(), self.defaultConfig)
        self.geolite_workspace = GeoLiteWorkspace(Path(__file__).resolve().parent / "Geolite")
        self.finding_validator_engine = FindingValidatorEngine()

        # Set up logging
        if loggingQueue is None:
            self.loggingQueue = mp.Queue()
            logListenerSetup(self.loggingQueue, self.config)
        else:
            self.loggingQueue = loggingQueue
        logWorkerSetup(self.loggingQueue)
        self.log = logging.getLogger(f"spiderfoot.{__name__}")

        cherrypy.config.update({
            'error_page.401': self.error_page_401,
            'error_page.404': self.error_page_404,
            'request.error_response': self.error_page
        })

        csp = (
            secure.ContentSecurityPolicy()
            .default_src("'self'")
            .script_src("'self'", "'unsafe-inline'", "blob:")
            .style_src("'self'", "'unsafe-inline'")
            .base_uri("'self'")
            .connect_src("'self'", "data:")
            .frame_src("'self'", 'data:')
            .img_src("'self'", "data:")
        )

        secure_headers = secure.Secure(
            server=secure.Server().set("server"),
            cache=secure.CacheControl().must_revalidate(),
            csp=csp,
            referrer=secure.ReferrerPolicy().no_referrer(),
        )

        cherrypy.config.update({
            "tools.response_headers.on": True,
            "tools.response_headers.headers": secure_headers.framework.cherrypy()
        })

    def error_page(self: 'SpiderFootWebUi') -> None:
        """Error page."""
        cherrypy.response.status = 500

        if self.config.get('_debug'):
            cherrypy.response.body = _cperror.get_error_page(status=500, traceback=_cperror.format_exc())
        else:
            cherrypy.response.body = b"<html><body>Error</body></html>"

    def error_page_401(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Unauthorized access HTTP 401 error page.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTML response
        """
        return ""

    def error_page_404(self: 'SpiderFootWebUi', status: str, message: str, traceback: str, version: str) -> str:
        """Not found error page 404.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTTP response template
        """
        templ = Template(filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message='Not Found', docroot=self.docroot, status=status, version=__version__)

    def jsonify_error(self: 'SpiderFootWebUi', status: str, message: str) -> dict:
        """Jsonify error response.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message

        Returns:
            dict: HTTP error response template
        """
        cherrypy.response.headers['Content-Type'] = 'application/json'
        cherrypy.response.status = status
        return {
            'error': {
                'http_status': status,
                'message': message,
            }
        }

    def error(self: 'SpiderFootWebUi', message: str) -> None:
        """Show generic error page with error message.

        Args:
            message (str): error message

        Returns:
            None
        """
        templ = Template(filename='spiderfoot/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message=message, docroot=self.docroot, version=__version__)

    def cleanUserInput(self: 'SpiderFootWebUi', inputList: list) -> list:
        """Convert data to HTML entities; except quotes and ampersands.

        Args:
            inputList (list): list of strings to sanitize

        Returns:
            list: sanitized input

        Raises:
            TypeError: inputList type was invalid

        Todo:
            Review all uses of this function, then remove it.
            Use of this function is overloaded.
        """
        if not isinstance(inputList, list):
            raise TypeError(f"inputList is {type(inputList)}; expected list()")

        ret = list()

        for item in inputList:
            if not item:
                ret.append('')
                continue
            c = html.escape(item, True)

            # Decode '&' and '"' HTML entities
            c = c.replace("&amp;", "&").replace("&quot;", "\"")
            ret.append(c)

        return ret

    def finding_status_labels(self: 'SpiderFootWebUi') -> dict:
        return {
            "novo": "Novo",
            "em_triagem": "Em triagem",
            "relevante": "Relevante",
            "descartado": "Descartado",
            "falso_positivo": "Falso positivo",
            "confirmado": "Confirmado"
        }

    def finding_relevance_labels(self: 'SpiderFootWebUi') -> dict:
        return {
            "pendente": "Pendente",
            "baixa": "Baixa",
            "media": "Média",
            "alta": "Alta",
            "critica": "Crítica"
        }

    def finding_exploitability_labels(self: 'SpiderFootWebUi') -> dict:
        return {
            "nao_avaliado": "Não avaliado",
            "nao_aplicavel": "Não aplicável",
            "nao_exploravel": "Não explorável",
            "potencial": "Potencial",
            "confirmada": "Confirmada"
        }

    def analyst_verdict_labels(self: 'SpiderFootWebUi') -> dict:
        return {
            "em_analise": "Em análise",
            "monitorar": "Monitorar",
            "sem_risco_pratico": "Sem risco prático",
            "exposto_confirmado": "Exposto e confirmado",
            "exploravel": "Explorável",
            "critico": "Crítico"
        }

    def _http_probe(self: 'SpiderFootWebUi', target: str, timeout: float = 4.0) -> dict:
        candidate_urls = []
        if re.match(r"^https?://", target, re.IGNORECASE):
            candidate_urls.append(target)
        else:
            candidate_urls.append(f"https://{target}")
            candidate_urls.append(f"http://{target}")

        last_error = None
        for url in candidate_urls:
            req = urllib.request.Request(url, method="GET", headers={"User-Agent": "SpiderFoot-Validation/1.0"})
            try:
                with urllib.request.urlopen(req, timeout=timeout, context=ssl.create_default_context()) as response:
                    return {
                        "status": "ok",
                        "url": url,
                        "code": getattr(response, "status", response.getcode()),
                        "server": response.headers.get("Server", ""),
                        "content_type": response.headers.get("Content-Type", "")
                    }
            except Exception as e:
                last_error = str(e)

        return {"status": "error", "error": last_error or "Não foi possível consultar o alvo via HTTP/HTTPS."}

    def _tcp_probe(self: 'SpiderFootWebUi', host: str, ports: list, timeout: float = 1.5) -> list:
        findings = []
        for port in ports:
            sock = socket.socket(socket.AF_INET6 if ":" in host else socket.AF_INET, socket.SOCK_STREAM)
            sock.settimeout(timeout)
            try:
                sock.connect((host, port))
                findings.append(port)
            except Exception:
                pass
            finally:
                sock.close()
        return findings

    def _run_finding_validation(self: 'SpiderFootWebUi', event_type: str, event_data: str) -> dict:
        return self.finding_validator_engine.validate(event_type, event_data)

        event_type = (event_type or "").upper()
        event_data = (event_data or "").strip()

        if not event_data:
            return {
                "validator": "sem_dados",
                "status": "erro",
                "summary": "O achado não possui dados suficientes para validação.",
                "details": "O valor do achado está vazio."
            }

        if event_type in ["IP_ADDRESS", "IPV6_ADDRESS"]:
            open_ports = self._tcp_probe(event_data, [22, 80, 443, 8080, 8443])
            reverse_name = ""
            try:
                reverse_name = socket.gethostbyaddr(event_data)[0]
            except Exception:
                reverse_name = ""

            summary = "IP respondeu em portas comuns." if open_ports else "IP sem resposta nas portas comuns testadas."
            details = {
                "ip": event_data,
                "reverse_dns": reverse_name,
                "open_ports": open_ports
            }
            return {"validator": "ip_basic_probe", "status": "ok" if open_ports else "warning", "summary": summary, "details": json.dumps(details, ensure_ascii=False, indent=2)}

        if event_type in ["INTERNET_NAME", "DOMAIN_NAME", "AFFILIATE_INTERNET_NAME", "AFFILIATE_DOMAIN_NAME", "SOCIAL_MEDIA", "ACCOUNT_EXTERNAL_OWNED", "SIMILAR_ACCOUNT_EXTERNAL"]:
            match = re.search(r"https?://[^\s>]+", event_data, re.IGNORECASE)
            probe_target = match.group(0) if match else event_data.split()[0]
            http_result = self._http_probe(probe_target)
            host = urllib.parse.urlparse(probe_target).hostname if re.match(r"^https?://", probe_target, re.IGNORECASE) else probe_target
            dns_records = []
            try:
                dns_records = sorted({item[4][0] for item in socket.getaddrinfo(host, None)})
            except Exception:
                dns_records = []

            ok = http_result.get("status") == "ok" or bool(dns_records)
            summary = "Hostname/URL validado com resposta observável." if ok else "Sem resposta conclusiva para hostname/URL."
            details = {
                "target": probe_target,
                "dns": dns_records,
                "http": http_result
            }
            return {"validator": "http_dns_probe", "status": "ok" if ok else "warning", "summary": summary, "details": json.dumps(details, ensure_ascii=False, indent=2)}

        if event_type == "EMAILADDR":
            if "@" not in event_data:
                return {"validator": "email_sanity", "status": "error", "summary": "E-mail malformado.", "details": event_data}
            domain = event_data.split("@", 1)[1]
            mx_probe = []
            try:
                mx_probe = sorted({item[4][0] for item in socket.getaddrinfo(domain, 25)})
            except Exception:
                mx_probe = []
            summary = "Domínio do e-mail possui resolução." if mx_probe else "Não foi possível confirmar resolução do domínio do e-mail."
            details = {"email": event_data, "domain": domain, "mx_like_resolution": mx_probe}
            return {"validator": "email_domain_probe", "status": "ok" if mx_probe else "warning", "summary": summary, "details": json.dumps(details, ensure_ascii=False, indent=2)}

        if event_type == "USERNAME":
            looks_valid = bool(re.match(r"^[A-Za-z0-9_.-]{3,64}$", event_data))
            summary = "Username possui formato consistente para pivot e validação manual." if looks_valid else "Username com formato incomum."
            details = {"username": event_data, "format_valid": looks_valid}
            return {"validator": "username_sanity", "status": "ok" if looks_valid else "warning", "summary": summary, "details": json.dumps(details, ensure_ascii=False, indent=2)}

        if event_type == "HUMAN_NAME":
            looks_valid = len(event_data.split()) >= 2
            summary = "Nome humano com formato plausível para investigação." if looks_valid else "Nome pouco conclusivo para validação automática."
            details = {"name": event_data, "word_count": len(event_data.split())}
            return {"validator": "human_name_sanity", "status": "ok" if looks_valid else "warning", "summary": summary, "details": json.dumps(details, ensure_ascii=False, indent=2)}

        return {
            "validator": "generic_observation",
            "status": "warning",
            "summary": "Ainda não há um validador automático específico para este tipo de achado.",
            "details": json.dumps({"event_type": event_type, "event_data": event_data}, ensure_ascii=False, indent=2)
        }

    def normalize_scan_target(self: 'SpiderFootWebUi', scantarget: str) -> tuple:
        """Normalize scan target input and infer quoted target types when obvious."""
        normalized_target = (scantarget or "").strip()
        target_type = SpiderFootHelpers.targetTypeFromString(normalized_target)

        if target_type is not None:
            return normalized_target, target_type

        if normalized_target.startswith('"') and normalized_target.endswith('"'):
            return normalized_target, None

        if " " in normalized_target:
            quoted_target = f'"{normalized_target}"'
            target_type = SpiderFootHelpers.targetTypeFromString(quoted_target)
            if target_type == "HUMAN_NAME":
                return quoted_target, target_type

        if re.match(r"^[A-Za-z0-9_.-]+$", normalized_target):
            quoted_target = f'"{normalized_target}"'
            target_type = SpiderFootHelpers.targetTypeFromString(quoted_target)
            if target_type == "USERNAME":
                return quoted_target, target_type

        return normalized_target, None

    def scan_presets(self: 'SpiderFootWebUi') -> list:
        """Return opinionated scan presets for common investigation workflows."""
        target_type_labels = {
            "INTERNET_NAME": "Domínio / Hostname / Subdomínio",
            "IP_ADDRESS": "IPv4",
            "IPV6_ADDRESS": "IPv6",
            "NETBLOCK_OWNER": "Sub-rede IPv4",
            "NETBLOCKV6_OWNER": "Sub-rede IPv6",
            "BGP_AS_OWNER": "ASN",
            "EMAILADDR": "E-mail",
            "HUMAN_NAME": "Nome de pessoa",
            "USERNAME": "Nome de usuário",
        }
        presets = [
            {
                "id": "preset_passive_domain_free",
                "name": "Domínio Passivo Sem Custo",
                "target_types": ["INTERNET_NAME"],
                "description": (
                    "Ideal para domínio, hostname e subdomínio. Prioriza fontes gratuitas "
                    "e módulos nativos para enumeração passiva, histórico, DNS, conteúdo e "
                    "infraestrutura relacionada, com baixo risco de interação direta com o alvo."
                ),
                "best_for": "Domínios, hostnames e subdomínios.",
                "caution": "Não é o melhor preset para IP isolado, telefone, e-mail ou Bitcoin.",
                "modules": [
                    "sfp_crt", "sfp_archiveorg", "sfp_commoncrawl", "sfp_github",
                    "sfp_grep_app", "sfp_dnsraw", "sfp_dnsresolve", "sfp_dnscommonsrv",
                    "sfp_dnsbrute", "sfp_robtex", "sfp_hackertarget", "sfp_threatcrowd",
                    "sfp_threatminer", "sfp_urlscan", "sfp_leakix", "sfp_whois",
                    "sfp_sslcert", "sfp_webserver", "sfp_webframework", "sfp_pageinfo",
                    "sfp_similar"
                ]
            },
            {
                "id": "preset_domain_investigate_balanced",
                "name": "Domínio Investigação Balanceada",
                "target_types": ["INTERNET_NAME"],
                "description": (
                    "Para investigar domínios com foco em sinais de abuso, phishing, exposição "
                    "e infraestrutura associada. Continua econômico, mas adiciona fontes de "
                    "reputação gratuitas para uma visão mais investigativa."
                ),
                "best_for": "Domínios suspeitos, phishing, IOC de hostname e investigação OSINT.",
                "caution": "Pode gerar mais ruído que o preset passivo puro.",
                "modules": [
                    "sfp_crt", "sfp_archiveorg", "sfp_commoncrawl", "sfp_github",
                    "sfp_grep_app", "sfp_dnsraw", "sfp_dnsresolve", "sfp_dnscommonsrv",
                    "sfp_dnsbrute", "sfp_robtex", "sfp_hackertarget", "sfp_threatcrowd",
                    "sfp_threatminer", "sfp_urlscan", "sfp_leakix", "sfp_openphish",
                    "sfp_phishtank", "sfp_botvrij", "sfp_stevenblack_hosts", "sfp_whois",
                    "sfp_sslcert", "sfp_webserver"
                ]
            },
            {
                "id": "preset_passive_ip_free",
                "name": "IP/Sub-rede Passivo Sem Custo",
                "target_types": ["IP_ADDRESS", "IPV6_ADDRESS", "NETBLOCK_OWNER", "NETBLOCKV6_OWNER", "BGP_AS_OWNER"],
                "description": (
                    "Voltado para IP, IPv6, sub-rede e ASN. Reúne RIR/BGP, reputação gratuita, "
                    "blacklists, passive DNS e vizinhança de rede para entender exposição e contexto."
                ),
                "best_for": "IP isolado, bloco CIDR, IPv6 e ASN.",
                "caution": "Para blocos grandes, prefira a aba GeoIP para filtrar e quebrar a rede antes.",
                "modules": [
                    "sfp_alienvaultiprep", "sfp_bgpview", "sfp_arin", "sfp_ripe",
                    "sfp_spamhaus", "sfp_spamcop", "sfp_sorbs", "sfp_blocklistde",
                    "sfp_threatfox", "sfp_threatcrowd", "sfp_threatminer", "sfp_isc",
                    "sfp_maltiverse", "sfp_leakix", "sfp_robtex", "sfp_hackertarget",
                    "sfp_dnsneighbor", "sfp_hosting"
                ]
            },
            {
                "id": "preset_breach_identity_free",
                "name": "Identidade e Vazamentos Sem Custo",
                "target_types": ["EMAILADDR", "INTERNET_NAME", "HUMAN_NAME", "USERNAME"],
                "description": (
                    "Focado em e-mail, domínio, nome e username para descobrir contas, menções, "
                    "indícios de vazamento e presença pública usando módulos gratuitos e internos."
                ),
                "best_for": "E-mail, domínio associado a pessoas, usernames e investigação de identidade.",
                "caution": "Não é um preset de infraestrutura; use os presets de domínio/IP para isso.",
                "modules": [
                    "sfp_emailformat", "sfp_skymem", "sfp_psbdmp", "sfp_wikileaks",
                    "sfp_accounts", "sfp_social", "sfp_pgp", "sfp_hunter",
                    "sfp_github", "sfp_grep_app"
                ]
            }
        ]

        available_modules = self.available_scan_modules()
        for preset in presets:
            preset["modules"] = [mod for mod in preset["modules"] if mod in available_modules]
            preset["module_count"] = len(preset["modules"])
            preset["target_type_labels"] = [target_type_labels.get(t, t) for t in preset["target_types"]]
        return presets

    def scan_preset_by_id(self: 'SpiderFootWebUi', preset_id: str) -> dict:
        for preset in self.scan_presets():
            if preset["id"] == preset_id:
                return preset
        return None

    def module_api_status(self: 'SpiderFootWebUi') -> dict:
        """Return availability status for each module based on in-app configuration."""
        statuses = dict()
        modules = self.config.get('__modules__', {})

        for module_name, module_info in modules.items():
            meta = module_info.get('meta') or {}
            flags = meta.get('flags') or []
            opts = module_info.get('opts') or {}
            optdescs = module_info.get('optdescs') or {}
            api_fields = [opt for opt in opts.keys() if 'api_key' in opt.lower()]
            api_required = 'apikey' in flags
            optional_api = False

            for opt_name in api_fields:
                description = str(optdescs.get(opt_name, '')).lower()
                if any(token in description for token in ['optional', 'without this', 'public api', 'limited to']):
                    optional_api = True
                    break

            configured = True
            if api_required and api_fields:
                configured = all(str(opts.get(field, '')).strip() != '' for field in api_fields)

            available = True
            reason = ""
            if api_required and not optional_api and api_fields and not configured:
                available = False
                reason = "Requer API configurada."
            elif api_required and optional_api and not configured:
                reason = "API não configurada, mas o módulo pode funcionar de forma limitada."
            elif api_required and configured:
                reason = "API configurada."

            statuses[module_name] = {
                "api_required": api_required,
                "api_optional": optional_api,
                "configured": configured,
                "available": available,
                "api_fields": api_fields,
                "reason": reason,
            }

        return statuses

    def available_scan_modules(self: 'SpiderFootWebUi') -> set:
        statuses = self.module_api_status()
        return {name for name, status in statuses.items() if status.get("available", True)}

    def type_availability(self: 'SpiderFootWebUi', event_types: list) -> dict:
        """Return whether each type has at least one currently available producer module."""
        cfg = deepcopy(self.config)
        sf = SpiderFoot(cfg)
        available_modules = self.available_scan_modules()
        module_definitions = self.config.get('__modules__', {})
        result = dict()

        for event_type in event_types:
            type_id = event_type[1]
            producers = sf.modulesProducing([type_id])
            enabled_producers = [mod for mod in producers if mod in available_modules]
            producer_details = []

            for producer in producers:
                producer_info = module_definitions.get(producer, {})
                producer_details.append({
                    "id": producer,
                    "name": producer_info.get("name", producer),
                    "available": producer in available_modules,
                })

            result[type_id] = {
                "available": len(enabled_producers) > 0,
                "producer_count": len(enabled_producers),
                "producers": producer_details,
                "reason": "" if enabled_producers else "Nenhum módulo disponível atualmente produz este tipo de dado.",
            }

        return result

    def scan_progress_summary(self: 'SpiderFootWebUi', scan_id: str) -> dict:
        """Build an execution progress summary for a scan."""
        dbh = SpiderFootDb(self.config)
        scan_info = dbh.scanInstanceGet(scan_id)
        if not scan_info:
            return {}

        scan_config = dbh.scanConfigGet(scan_id) or {}
        enabled_modules = scan_config.get('_modulesenabled', '')
        if isinstance(enabled_modules, str):
            enabled_modules = [m for m in enabled_modules.split(',') if m]
        elif not isinstance(enabled_modules, list):
            enabled_modules = []

        enabled_modules = [
            mod for mod in enabled_modules
            if mod not in ["sfp__stor_db", "sfp__stor_stdout"]
        ]

        module_names = {
            mod: (self.config.get('__modules__', {}).get(mod, {}) or {}).get('name', mod)
            for mod in enabled_modules
        }

        module_summary = dbh.scanResultSummary(scan_id, by="module") or []
        modules_with_results = []
        module_result_counts = {}
        for row in module_summary:
            module_id = row[0]
            if not module_id or module_id == "SpiderFoot UI":
                continue
            modules_with_results.append(module_id)
            module_result_counts[module_id] = {
                "total": row[3],
                "unique": row[4],
                "last_seen": row[2],
            }

        modules_with_results = sorted(set(modules_with_results))

        recent_logs = dbh.scanLogs(scan_id, limit=100) or []
        recent_logs = list(reversed(recent_logs))
        latest_log_message = ""
        queue_total = None
        queue_preview = ""
        running_modules = []
        errored_modules = []
        loaded_modules = []

        for log_row in recent_logs:
            message = str(log_row[3] or "")
            if not latest_log_message and message:
                latest_log_message = message

            loaded_match = re.search(r"^([a-zA-Z0-9_]+) module loaded\.$", message)
            if loaded_match:
                loaded_modules.append(loaded_match.group(1))

            running_match = re.search(r"^Modules running: \d+ \((.*)\)$", message)
            if running_match:
                running_modules = [m.strip() for m in running_match.group(1).split(',') if m.strip()]

            errored_match = re.search(r"^Modules errored: \d+ \((.*)\)$", message)
            if errored_match:
                errored_modules = [m.strip() for m in errored_match.group(1).split(',') if m.strip()]

            queued_match = re.search(r"^Events queued: ([0-9,]+) \((.*)\)$", message)
            if queued_match:
                queue_total = int(queued_match.group(1).replace(',', ''))
                queue_preview = queued_match.group(2).strip()

        loaded_modules = sorted(set(loaded_modules))
        errored_modules = sorted(set(errored_modules))

        total_modules = len(enabled_modules)
        active_modules = len([m for m in modules_with_results if m in enabled_modules])
        loaded_count = len([m for m in loaded_modules if m in enabled_modules]) or total_modules
        progress_base = loaded_count or total_modules or 1
        progress_percent = int(round((active_modules / progress_base) * 100)) if progress_base else 0
        progress_percent = max(0, min(progress_percent, 100))

        pending_modules = [m for m in enabled_modules if m not in modules_with_results]
        if running_modules:
            pending_modules = [m for m in pending_modules if m not in running_modules]

        top_active_modules = sorted(
            [m for m in modules_with_results if m in enabled_modules],
            key=lambda mod: module_result_counts.get(mod, {}).get("total", 0),
            reverse=True
        )[:8]

        return {
            "status": scan_info[5],
            "total_modules": total_modules,
            "loaded_modules": len([m for m in loaded_modules if m in enabled_modules]) or total_modules,
            "active_modules": active_modules,
            "pending_modules": len(pending_modules),
            "running_modules": running_modules,
            "errored_modules": errored_modules,
            "queue_total": queue_total,
            "queue_preview": queue_preview,
            "progress_percent": progress_percent,
            "latest_log_message": latest_log_message,
            "top_active_modules": [
                {
                    "id": mod,
                    "name": module_names.get(mod, mod),
                    "total": module_result_counts.get(mod, {}).get("total", 0),
                    "unique": module_result_counts.get(mod, {}).get("unique", 0),
                }
                for mod in top_active_modules
            ],
            "pending_module_names": [module_names.get(mod, mod) for mod in pending_modules[:8]],
            "running_module_names": [module_names.get(mod, mod) for mod in running_modules[:8]],
            "errored_module_names": [module_names.get(mod, mod) for mod in errored_modules[:8]],
        }

    def searchBase(self: 'SpiderFootWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search.

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD

        Returns:
            list: search results
        """
        retdata = []

        if not id and not eventType and not value:
            return retdata

        if not value:
            value = ''

        regex = ""
        if value.startswith("/") and value.endswith("/"):
            regex = value[1:len(value) - 1]
            value = ""

        value = value.replace('*', '%')
        if value in [None, ""] and regex in [None, ""]:
            value = "%"
            regex = ""

        dbh = SpiderFootDb(self.config)
        criteria = {
            'scan_id': id or '',
            'type': eventType or '',
            'value': value or '',
            'regex': regex or '',
        }

        try:
            data = dbh.search(criteria)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            escapeddata = html.escape(row[1])
            escapedsrc = html.escape(row[2])
            retdata.append([lastseen, escapeddata, escapedsrc,
                            row[3], row[5], row[6], row[7], row[8], row[10],
                            row[11], row[4], row[13], row[14]])

        return retdata

    def buildExcel(self: 'SpiderFootWebUi', data: list, columnNames: list, sheetNameIndex: int = 0) -> str:
        """Convert supplied raw data into GEXF (Graph Exchange XML Format) format (e.g. for Gephi).

        Args:
            data (list): Scan result as list
            columnNames (list): column names
            sheetNameIndex (int): TBD

        Returns:
            str: Excel workbook
        """
        rowNums = dict()
        workbook = openpyxl.Workbook()
        defaultSheet = workbook.active
        columnNames.pop(sheetNameIndex)
        allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'
        for row in data:
            sheetName = "".join([c for c in str(row.pop(sheetNameIndex)) if c.upper() in allowed_sheet_chars])
            try:
                sheet = workbook[sheetName]
            except KeyError:
                # Create sheet
                workbook.create_sheet(sheetName)
                sheet = workbook[sheetName]
                # Write headers
                for col_num, column_title in enumerate(columnNames, 1):
                    cell = sheet.cell(row=1, column=col_num)
                    cell.value = column_title
                rowNums[sheetName] = 2

            # Write row
            for col_num, cell_value in enumerate(row, 1):
                cell = sheet.cell(row=rowNums[sheetName], column=col_num)
                cell.value = cell_value

            rowNums[sheetName] += 1

        if rowNums:
            workbook.remove(defaultSheet)

        # Sort sheets alphabetically
        workbook._sheets.sort(key=lambda ws: ws.title)

        # Save workbook
        with BytesIO() as f:
            workbook.save(f)
            f.seek(0)
            return f.read()

    #
    # USER INTERFACE PAGES
    #

    @cherrypy.expose
    def scanexportlogs(self: 'SpiderFootWebUi', id: str, dialect: str = "excel") -> bytes:
        """Get scan log

        Args:
            id (str): scan ID
            dialect (str): CSV dialect (default: excel)

        Returns:
            bytes: scan logs in CSV format
        """
        dbh = SpiderFootDb(self.config)

        try:
            data = dbh.scanLogs(id, None, None, True)
        except Exception:
            return self.error("Scan ID not found.")

        if not data:
            return self.error("Scan ID not found.")

        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(["Date", "Component", "Type", "Event", "Event ID"])
        for row in data:
            parser.writerow([
                time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000)),
                str(row[1]),
                str(row[2]),
                str(row[3]),
                row[4]
            ])

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename=SpiderFoot-{id}.log.csv"
        cherrypy.response.headers['Content-Type'] = "application/csv"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return fileobj.getvalue().encode('utf-8')

    @cherrypy.expose
    def scancorrelationsexport(self: 'SpiderFootWebUi', id: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan correlation data in CSV or Excel format.

        Args:
            id (str): scan ID
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)

        try:
            scaninfo = dbh.scanInstanceGet(id)
            scan_name = scaninfo[0]
        except Exception:
            return json.dumps(["ERROR", "Could not retrieve info for scan."]).encode('utf-8')

        try:
            correlations = dbh.scanCorrelationList(id)
        except Exception:
            return json.dumps(["ERROR", "Could not retrieve correlations for scan."]).encode('utf-8')

        headings = ["Rule Name", "Correlation", "Risk", "Description"]

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in correlations:
                correlation = row[1]
                rule_name = row[2]
                rule_risk = row[3]
                rule_description = row[5]
                rows.append([rule_name, correlation, rule_risk, rule_description])

            if scan_name:
                fname = f"{scan_name}-SpiderFoot-correlations.xlxs"
            else:
                fname = "SpiderFoot-correlations.xlxs"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, headings, sheetNameIndex=0)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(headings)

            for row in correlations:
                correlation = row[1]
                rule_name = row[2]
                rule_risk = row[3]
                rule_description = row[5]
                parser.writerow([rule_name, correlation, rule_risk, rule_description])

            if scan_name:
                fname = f"{scan_name}-SpiderFoot-correlations.csv"
            else:
                fname = "SpiderFoot-correlations.csv"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scaneventresultexport(self: 'SpiderFootWebUi', id: str, type: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format

        Args:
            id (str): scan ID
            type (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanResultEvent(id, type)

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

            fname = "SpiderFoot.xlsx"
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

            fname = "SpiderFoot.csv"
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scaneventresultexportmulti(self: 'SpiderFootWebUi', ids: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format for multiple scans

        Args:
            ids (str): comma separated list of scan IDs
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = SpiderFootDb(self.config)
        scaninfo = dict()
        data = list()
        scan_name = ""

        for id in ids.split(','):
            scaninfo[id] = dbh.scanInstanceGet(id)
            if scaninfo[id] is None:
                continue
            scan_name = scaninfo[id][0]
            data = data + dbh.scanResultEvent(id)

        if not data:
            return None

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]),
                            str(row[2]), row[13], datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "SpiderFoot.xlsx"
            else:
                fname = scan_name + "-SpiderFoot.xlsx"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Scan Name", "Updated", "Type", "Module",
                                   "Source", "F/P", "Data"], sheetNameIndex=2)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Scan Name", "Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]),
                                str(row[2]), row[13], datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "SpiderFoot.csv"
            else:
                fname = scan_name + "-SpiderFoot.csv"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scansearchresultexport(self: 'SpiderFootWebUi', id: str, eventType: str = None, value: str = None, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get search result data in CSV or Excel format

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        data = self.searchBase(id, eventType, value)

        if not data:
            return None

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=SpiderFoot.xlsx"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["Updated", "Type", "Module", "Source",
                                   "F/P", "Data"], sheetNameIndex=1)

        if filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["Updated", "Type", "Module", "Source", "F/P", "Data"])
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=SpiderFoot.csv"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scanexportjsonmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        """Get scan event result data in JSON format for multiple scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: results in JSON format
        """
        dbh = SpiderFootDb(self.config)
        scaninfo = list()
        scan_name = ""

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)

            if scan is None:
                continue

            scan_name = scan[0]

            for row in dbh.scanResultEvent(id):
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                event_data = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                source_data = str(row[2])
                source_module = str(row[3])
                event_type = row[4]
                false_positive = row[13]

                if event_type == "ROOT":
                    continue

                scaninfo.append({
                    "data": event_data,
                    "event_type": event_type,
                    "module": source_module,
                    "source_data": source_data,
                    "false_positive": false_positive,
                    "last_seen": lastseen,
                    "scan_name": scan_name,
                    "scan_target": scan[1]
                })

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "SpiderFoot.json"
        else:
            fname = scan_name + "-SpiderFoot.json"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return json.dumps(scaninfo).encode('utf-8')

    @cherrypy.expose
    def scanviz(self: 'SpiderFootWebUi', id: str, gexf: str = "0") -> str:
        """Export entities from scan results for visualising.

        Args:
            id (str): scan ID
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        if not id:
            return None

        dbh = SpiderFootDb(self.config)
        data = dbh.scanResultEvent(id, filterFp=True)
        scan = dbh.scanInstanceGet(id)

        if not scan:
            return None

        scan_name = scan[0]

        root = scan[1]

        if gexf == "0":
            return SpiderFootHelpers.buildGraphJson([root], data)

        if not scan_name:
            fname = "SpiderFoot.gexf"
        else:
            fname = scan_name + "SpiderFoot.gexf"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return SpiderFootHelpers.buildGraphGexf([root], "SpiderFoot Export", data)

    @cherrypy.expose
    def scanvizmulti(self: 'SpiderFootWebUi', ids: str, gexf: str = "1") -> str:
        """Export entities results from multiple scans in GEXF format.

        Args:
            ids (str): scan IDs
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        dbh = SpiderFootDb(self.config)
        data = list()
        roots = list()
        scan_name = ""

        if not ids:
            return None

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)
            if not scan:
                continue
            data = data + dbh.scanResultEvent(id, filterFp=True)
            roots.append(scan[1])
            scan_name = scan[0]

        if not data:
            return None

        if gexf == "0":
            # Not implemented yet
            return None

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "SpiderFoot.gexf"
        else:
            fname = scan_name + "-SpiderFoot.gexf"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return SpiderFootHelpers.buildGraphGexf(roots, "SpiderFoot Export", data)

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanopts(self: 'SpiderFootWebUi', id: str) -> dict:
        """Return configuration used for the specified scan as JSON.

        Args:
            id: scan ID

        Returns:
            dict: scan options for the specified scan
        """
        dbh = SpiderFootDb(self.config)
        ret = dict()

        meta = dbh.scanInstanceGet(id)
        if not meta:
            return ret

        if meta[3] != 0:
            started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[3]))
        else:
            started = "Not yet"

        if meta[4] != 0:
            finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[4]))
        else:
            finished = "Not yet"

        ret['meta'] = [meta[0], meta[1], meta[2], started, finished, meta[5]]
        ret['config'] = dbh.scanConfigGet(id)
        ret['configdesc'] = dict()
        for key in list(ret['config'].keys()):
            if ':' not in key:
                globaloptdescs = self.config['__globaloptdescs__']
                if globaloptdescs:
                    ret['configdesc'][key] = globaloptdescs.get(key, f"{key} (legacy)")
            else:
                [modName, modOpt] = key.split(':')
                if modName not in list(self.config['__modules__'].keys()):
                    continue

                if modOpt not in list(self.config['__modules__'][modName]['optdescs'].keys()):
                    continue

                ret['configdesc'][key] = self.config['__modules__'][modName]['optdescs'][modOpt]

        return ret

    @cherrypy.expose
    def rerunscan(self: 'SpiderFootWebUi', id: str) -> None:
        """Rerun a scan.

        Args:
            id (str): scan ID

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to info page for new scan
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)
        info = dbh.scanInstanceGet(id)

        if not info:
            return self.error("Invalid scan ID.")

        scanname = info[0]
        scantarget = info[1]

        scanconfig = dbh.scanConfigGet(id)
        if not scanconfig:
            return self.error(f"Error loading config from scan: {id}")

        modlist = scanconfig['_modulesenabled'].split(',')
        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        scantarget, targetType = self.normalize_scan_target(scantarget)
        if not targetType:
            return self.error("Invalid target type. Could not recognize it as a target SpiderFoot supports.")

        if targetType not in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.lower()
        else:
            scantarget = scantarget.replace("\"", "")

        # Start running a new scan
        scanId = SpiderFootHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}")
            return self.error(f"[-] Scan [{scanId}] failed: {e}")

        # Wait until the scan has initialized
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={scanId}", status=302)

    @cherrypy.expose
    def rerunscanmulti(self: 'SpiderFootWebUi', ids: str) -> str:
        """Rerun scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: Scan list page HTML
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = SpiderFootDb(cfg)

        for id in ids.split(","):
            info = dbh.scanInstanceGet(id)
            if not info:
                return self.error("Invalid scan ID.")

            scanconfig = dbh.scanConfigGet(id)
            scanname = info[0]
            scantarget = info[1]
            targetType = None

            if len(scanconfig) == 0:
                return self.error("Something went wrong internally.")

            modlist = scanconfig['_modulesenabled'].split(',')
            if "sfp__stor_stdout" in modlist:
                modlist.remove("sfp__stor_stdout")

            scantarget, targetType = self.normalize_scan_target(scantarget)
            if targetType is None:
                return self.error("Invalid target type. Could not recognize it as a target SpiderFoot supports.")

            if targetType not in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
                scantarget = scantarget.lower()
            else:
                scantarget = scantarget.replace("\"", "")

            # Start running a new scan
            scanId = SpiderFootHelpers.genScanInstanceId()
            try:
                p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
                p.daemon = True
                p.start()
            except Exception as e:
                self.log.error(f"[-] Scan [{scanId}] failed: {e}")
                return self.error(f"[-] Scan [{scanId}] failed: {e}")

            # Wait until the scan has initialized
            while dbh.scanInstanceGet(scanId) is None:
                self.log.info("Waiting for the scan to initialize...")
                time.sleep(1)

        templ = Template(filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(rerunscans=True, docroot=self.docroot, pageid="SCANLIST", version=__version__)

    @cherrypy.expose
    def newscan(self: 'SpiderFootWebUi') -> str:
        """Configure a new scan.

        Returns:
            str: New scan page HTML
        """
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        module_status = self.module_api_status()
        type_status = self.type_availability(types)
        templ = Template(filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], scanname="",
                            selectedmods="", selectedpreset="", presets=self.scan_presets(),
                            module_status=module_status, type_status=type_status,
                            scantarget="", version=__version__)

    @cherrypy.expose
    def clonescan(self: 'SpiderFootWebUi', id: str) -> str:
        """Clone an existing scan (pre-selected options in the newscan page).

        Args:
            id (str): scan ID to clone

        Returns:
            str: New scan page HTML pre-populated with options from cloned scan.
        """
        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        module_status = self.module_api_status()
        type_status = self.type_availability(types)
        info = dbh.scanInstanceGet(id)

        if not info:
            return self.error("Invalid scan ID.")

        scanconfig = dbh.scanConfigGet(id)
        scanname = info[0]
        scantarget = info[1]
        targetType = None

        if scanname == "" or scantarget == "" or len(scanconfig) == 0:
            return self.error("Something went wrong internally.")

        targetType = SpiderFootHelpers.targetTypeFromString(scantarget)
        if targetType is None:
            # It must be a name, so wrap quotes around it
            scantarget = "&quot;" + scantarget + "&quot;"

        modlist = scanconfig['_modulesenabled'].split(',')

        templ = Template(filename='spiderfoot/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], selectedmods=modlist,
                            selectedpreset="",
                            presets=self.scan_presets(),
                            module_status=module_status, type_status=type_status,
                            scanname=str(scanname),
                            scantarget=str(scantarget), version=__version__)

    @cherrypy.expose
    def index(self: 'SpiderFootWebUi') -> str:
        """Show scan list page.

        Returns:
            str: Scan list page HTML
        """
        templ = Template(filename='spiderfoot/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(pageid='SCANLIST', docroot=self.docroot, version=__version__)

    @cherrypy.expose
    def scaninfo(self: 'SpiderFootWebUi', id: str) -> str:
        """Information about a selected scan.

        Args:
            id (str): scan id

        Returns:
            str: scan info page HTML
        """
        dbh = SpiderFootDb(self.config)
        res = dbh.scanInstanceGet(id)
        if res is None:
            return self.error("Scan ID not found.")

        templ = Template(filename='spiderfoot/templates/scaninfo.tmpl', lookup=self.lookup, input_encoding='utf-8')
        return templ.render(id=id, name=html.escape(res[0]), status=res[5], docroot=self.docroot, version=__version__,
                            pageid="SCANLIST")

    @cherrypy.expose
    def opts(self: 'SpiderFootWebUi', updated: str = None, module: str = None) -> str:
        """Show module and global settings page.

        Args:
            updated (str): scan options were updated successfully
            module (str): module identifier to focus in the settings page

        Returns:
            str: scan options page HTML
        """
        templ = Template(filename='spiderfoot/templates/opts.tmpl', lookup=self.lookup)
        self.token = random.SystemRandom().randint(0, 99999999)
        return templ.render(opts=self.config, pageid='SETTINGS', token=self.token, version=__version__,
                            updated=updated, selected_module=module, docroot=self.docroot)

    @cherrypy.expose
    def optsexport(self: 'SpiderFootWebUi', pattern: str = None) -> str:
        """Export configuration.

        Args:
            pattern (str): TBD

        Returns:
            str: Configuration settings
        """
        sf = SpiderFoot(self.config)
        conf = sf.configSerialize(self.config)
        content = ""

        for opt in sorted(conf):
            if ":_" in opt or opt.startswith("_"):
                continue

            if pattern:
                if pattern in opt:
                    content += f"{opt}={conf[opt]}\n"
            else:
                content += f"{opt}={conf[opt]}\n"

        cherrypy.response.headers['Content-Disposition'] = 'attachment; filename="SpiderFoot.cfg"'
        cherrypy.response.headers['Content-Type'] = "text/plain"
        return content

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def optsraw(self: 'SpiderFootWebUi') -> str:
        """Return global and module settings as json.

        Returns:
            str: settings as JSON
        """
        ret = dict()
        self.token = random.SystemRandom().randint(0, 99999999)
        for opt in self.config:
            if not opt.startswith('__'):
                ret["global." + opt] = self.config[opt]
                continue

            if opt == '__modules__':
                for mod in sorted(self.config['__modules__'].keys()):
                    for mo in sorted(self.config['__modules__'][mod]['opts'].keys()):
                        if mo.startswith("_"):
                            continue
                        ret["module." + mod + "." + mo] = self.config['__modules__'][mod]['opts'][mo]

        return ['SUCCESS', {'token': self.token, 'data': ret}]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scandelete(self: 'SpiderFootWebUi', id: str) -> str:
        """Delete scan(s).

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            if res[5] in ["RUNNING", "STARTING", "STARTED"]:
                return self.jsonify_error('400', f"Scan {scan_id} is {res[5]}. You cannot delete running scans.")

        for scan_id in ids:
            dbh.scanInstanceDelete(scan_id)

        return ""

    @cherrypy.expose
    def savesettings(self: 'SpiderFootWebUi', allopts: str, token: str, configFile: 'cherrypy._cpreqbody.Part' = None) -> None:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token
            configFile (cherrypy._cpreqbody.Part): TBD

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to scan settings
        """
        if str(token) != str(self.token):
            return self.error(f"Invalid token ({token})")

        # configFile seems to get set even if a file isn't uploaded
        if configFile and configFile.file:
            try:
                contents = configFile.file.read()

                if isinstance(contents, bytes):
                    contents = contents.decode('utf-8')

                tmp = dict()
                for line in contents.split("\n"):
                    if "=" not in line:
                        continue

                    opt_array = line.strip().split("=")
                    if len(opt_array) == 1:
                        opt_array[1] = ""

                    tmp[opt_array[0]] = '='.join(opt_array[1:])

                allopts = json.dumps(tmp).encode('utf-8')
            except Exception as e:
                return self.error(f"Failed to parse input file. Was it generated from SpiderFoot? ({e})")

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")
            return self.error("Failed to reset settings")

        # Save settings
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return self.error(f"Processing one or more of your inputs failed: {e}")

        raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")

    @cherrypy.expose
    def savesettingsraw(self: 'SpiderFootWebUi', allopts: str, token: str) -> str:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token

        Returns:
            str: save success as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if str(token) != str(self.token):
            return json.dumps(["ERROR", f"Invalid token ({token})."]).encode('utf-8')

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Failed to reset settings"]).encode('utf-8')

        # Save settings
        try:
            dbh = SpiderFootDb(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            sf = SpiderFoot(self.config)
            self.config = sf.configUnserialize(cleanopts, currentopts)
            dbh.configSet(sf.configSerialize(self.config))
        except Exception as e:
            return json.dumps(["ERROR", f"Processing one or more of your inputs failed: {e}"]).encode('utf-8')

        return json.dumps(["SUCCESS", ""]).encode('utf-8')

    def reset_settings(self: 'SpiderFootWebUi') -> bool:
        """Reset settings to default.

        Returns:
            bool: success
        """
        try:
            dbh = SpiderFootDb(self.config)
            dbh.configClear()  # Clear it in the DB
            self.config = deepcopy(self.defaultConfig)  # Clear in memory
        except Exception:
            return False

        return True

    @cherrypy.expose
    def resultsetfp(self: 'SpiderFootWebUi', id: str, resultids: str, fp: str) -> str:
        """Set a bunch of results (hashes) as false positive.

        Args:
            id (str): scan ID
            resultids (str): comma separated list of result IDs
            fp (str): 0 or 1

        Returns:
            str: set false positive status as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)

        if fp not in ["0", "1"]:
            return json.dumps(["ERROR", "No FP flag set or not set correctly."]).encode('utf-8')

        try:
            ids = json.loads(resultids)
        except Exception:
            return json.dumps(["ERROR", "No IDs supplied."]).encode('utf-8')

        # Cannot set FPs if a scan is not completed
        status = dbh.scanInstanceGet(id)
        if not status:
            return self.error(f"Invalid scan ID: {id}")

        if status[5] not in ["ABORTED", "FINISHED", "ERROR-FAILED"]:
            return json.dumps([
                "WARNING",
                "Scan must be in a finished state when setting False Positives."
            ]).encode('utf-8')

        # Make sure the user doesn't set something as non-FP when the
        # parent is set as an FP.
        if fp == "0":
            data = dbh.scanElementSourcesDirect(id, ids)
            for row in data:
                if str(row[14]) == "1":
                    return json.dumps([
                        "WARNING",
                        f"Cannot unset element {id} as False Positive if a parent element is still False Positive."
                    ]).encode('utf-8')

        # Set all the children as FPs too.. it's only logical afterall, right?
        childs = dbh.scanElementChildrenAll(id, ids)
        allIds = ids + childs

        ret = dbh.scanResultsUpdateFP(id, allIds, fp)
        if ret:
            return json.dumps(["SUCCESS", ""]).encode('utf-8')

        return json.dumps(["ERROR", "Exception encountered."]).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def eventtypes(self: 'SpiderFootWebUi') -> list:
        """List all event types.

        Returns:
            list: list of event types
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = SpiderFootDb(self.config)
        types = dbh.eventTypes()
        ret = list()

        for r in types:
            ret.append([r[1], r[0]])

        return sorted(ret, key=itemgetter(0))

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def modules(self: 'SpiderFootWebUi') -> list:
        """List all modules.

        Returns:
            list: list of modules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        ret = list()

        modinfo = list(self.config['__modules__'].keys())
        if not modinfo:
            return ret

        modinfo.sort()

        for m in modinfo:
            if "__" in m:
                continue
            ret.append({'name': m, 'descr': self.config['__modules__'][m]['descr']})

        return ret

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def correlationrules(self: 'SpiderFootWebUi') -> list:
        """List all correlation rules.

        Returns:
            list: list of correlation rules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        ret = list()

        rules = self.config['__correlationrules__']
        if not rules:
            return ret

        for r in rules:
            ret.append({
                'id': r['id'],
                'name': r['meta']['name'],
                'descr': r['meta']['description'],
                'risk': r['meta']['risk'],
            })

        return ret

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def ping(self: 'SpiderFootWebUi') -> list:
        """For the CLI to test connectivity to this server.

        Returns:
            list: SpiderFoot version as JSON
        """
        return ["SUCCESS", __version__]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def query(self: 'SpiderFootWebUi', query: str) -> str:
        """For the CLI to run queries against the database.

        Args:
            query (str): SQL query

        Returns:
            str: query results as JSON
        """
        dbh = SpiderFootDb(self.config)

        if not query:
            return self.jsonify_error('400', "Invalid query.")

        if not query.lower().startswith("select"):
            return self.jsonify_error('400', "Non-SELECTs are unpredictable and not recommended.")

        try:
            ret = dbh.dbh.execute(query)
            data = ret.fetchall()
            columnNames = [c[0] for c in dbh.dbh.description]
            return [dict(zip(columnNames, row)) for row in data]
        except Exception as e:
            return self.jsonify_error('500', str(e))

    @cherrypy.expose
    def geoip(self: 'SpiderFootWebUi') -> str:
        """Show GeoIP bulk scan page."""
        templ = Template(filename='spiderfoot/templates/geoip.tmpl', lookup=self.lookup)
        return templ.render(pageid='GEOIP', docroot=self.docroot, version=__version__)

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def geoipdatasets(self: 'SpiderFootWebUi') -> dict:
        """List GeoLite datasets discovered under the Geolite directory."""
        try:
            return self.geolite_workspace.available_files()
        except Exception as e:
            return self.jsonify_error('500', str(e))

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def geoippreview(
        self: 'SpiderFootWebUi',
        city_blocks_file: str,
        city_locations_file: str = None,
        country_locations_file: str = None,
        asn_blocks_file: str = None,
        network_filter: str = None,
        city_filter: str = None,
        country_filter: str = None,
        organization_filter: str = None,
        asn_filter: str = None,
        limit: str = "1000",
    ) -> dict:
        """Preview merged GeoLite rows for selection in the web UI."""
        try:
            if not city_locations_file or not country_locations_file or not asn_blocks_file:
                return self.jsonify_error('400', "City Locations, Country Locations and ASN Blocks are required before expanding networks.")

            has_filter = any([
                network_filter,
                city_filter,
                country_filter,
                organization_filter,
                asn_filter,
            ])
            if not has_filter:
                return self.jsonify_error('400', "At least one filter is required before expanding GeoLite networks.")

            result = self.geolite_workspace.preview_rows(
                city_blocks_file=city_blocks_file,
                city_locations_file=city_locations_file or None,
                country_locations_file=country_locations_file or None,
                asn_blocks_file=asn_blocks_file or None,
                filters={
                    "network": network_filter,
                    "city": city_filter,
                    "country": country_filter,
                    "organization": organization_filter,
                    "asn": asn_filter,
                },
                limit=int(limit or 1000),
            )
            result["selected_files"] = {
                "city_blocks_file": city_blocks_file,
                "city_locations_file": city_locations_file or "",
                "country_locations_file": country_locations_file or "",
                "asn_blocks_file": asn_blocks_file or "",
            }
            return result
        except Exception as e:
            return self.jsonify_error('500', str(e))

    def _start_scan_instance(
        self: 'SpiderFootWebUi',
        scanname: str,
        scantarget: str,
        modulelist: str = "",
        typelist: str = "",
        usecase: str = "",
        preset: str = None,
    ) -> list:
        scanname = self.cleanUserInput([scanname])[0]
        scantarget = self.cleanUserInput([scantarget])[0]

        if not scanname:
            return ["ERROR", "Incorrect usage: scan name was not specified."]

        if not scantarget:
            return ["ERROR", "Incorrect usage: scan target was not specified."]

        if not typelist and not modulelist and not usecase and not preset:
            return ["ERROR", "Incorrect usage: no modules specified for scan."]

        scantarget, targetType = self.normalize_scan_target(scantarget)
        if targetType is None:
            return ["ERROR", "Unrecognised target type."]

        dbh = SpiderFootDb(self.config)
        cfg = deepcopy(self.config)
        sf = SpiderFoot(cfg)
        available_modules = self.available_scan_modules()

        modlist = list()
        preset_info = None

        if preset:
            preset_info = self.scan_preset_by_id(preset)
            if not preset_info:
                return ["ERROR", f"Unknown scan preset: {preset}"]

            if targetType not in preset_info["target_types"]:
                return ["ERROR", f"Preset '{preset_info['name']}' is not compatible with target type {targetType}."]

            modlist = deepcopy(preset_info["modules"])

        if not modlist and modulelist:
            modlist = modulelist.replace('module_', '').split(',')
            invalid_modules = [mod for mod in modlist if mod not in available_modules]
            if invalid_modules:
                return ["ERROR", f"These modules are unavailable because required API credentials are not configured: {', '.join(invalid_modules)}"]

        if len(modlist) == 0 and typelist:
            typesx = typelist.replace('type_', '').split(',')
            modlist = [mod for mod in sf.modulesProducing(typesx) if mod in available_modules]
            newmods = deepcopy(modlist)
            newmodcpy = deepcopy(newmods)

            while len(newmodcpy) > 0:
                for etype in sf.eventsToModules(newmodcpy):
                    xmods = [mod for mod in sf.modulesProducing([etype]) if mod in available_modules]
                    for mod in xmods:
                        if mod not in modlist:
                            modlist.append(mod)
                            newmods.append(mod)
                newmodcpy = deepcopy(newmods)
                newmods = list()

        if len(modlist) == 0 and usecase:
            for mod in self.config['__modules__']:
                if mod in available_modules and (usecase == 'all' or usecase in self.config['__modules__'][mod]['group']):
                    modlist.append(mod)

        if not modlist:
            return ["ERROR", "Incorrect usage: no modules specified for scan."]

        if "sfp__stor_db" not in modlist:
            modlist.append("sfp__stor_db")
        modlist.sort()

        if "sfp__stor_stdout" in modlist:
            modlist.remove("sfp__stor_stdout")

        if targetType in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.replace("\"", "")
        else:
            scantarget = scantarget.lower()

        scanId = SpiderFootHelpers.genScanInstanceId()
        try:
            p = mp.Process(target=startSpiderFootScanner, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] Scan [{scanId}] failed: {e}")
            return ["ERROR", f"[-] Scan [{scanId}] failed: {e}"]

        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("Waiting for the scan to initialize...")
            time.sleep(1)

        return ["SUCCESS", scanId]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def geoipstartscans(
        self: 'SpiderFootWebUi',
        selections: str,
        usecase: str = "all",
        max_ips_per_network: str = "0",
        usable_hosts_only: str = "1",
        name_template: str = "{ip} - AS{asn} - {org}",
        max_scans: str = "0",
    ) -> dict:
        """Expand selected GeoLite networks and start one scan per IP."""
        try:
            selected_rows = json.loads(selections or "[]")
        except json.JSONDecodeError:
            return self.jsonify_error('400', "Invalid GeoIP selection payload.")

        if not isinstance(selected_rows, list) or len(selected_rows) == 0:
            return self.jsonify_error('400', "No GeoIP rows selected.")

        try:
            max_ips = int(max_ips_per_network or 0)
            max_total_scans = int(max_scans or 0)
        except ValueError:
            return self.jsonify_error('400', "Invalid numeric GeoIP scan limits.")

        usable_only = str(usable_hosts_only).lower() in ["1", "true", "yes", "on"]

        created = []
        failures = []
        total_requested = 0

        for row in selected_rows:
            network_text = (row.get("network") or "").strip()
            if not network_text:
                continue

            try:
                network = ipaddress.ip_network(network_text, strict=False)
            except Exception as e:
                failures.append({"network": network_text, "error": str(e)})
                continue

            ip_iterable = network.hosts() if usable_only and getattr(network, "version", None) == 4 else network
            row_count = 0

            for ip in ip_iterable:
                if max_ips and row_count >= max_ips:
                    break
                if max_total_scans and total_requested >= max_total_scans:
                    break

                ip_text = str(ip)
                try:
                    scanname = name_template.format(
                        ip=ip_text,
                        network=network_text,
                        asn=(row.get("asn") or "SEM-ASN"),
                        org=(row.get("organization") or "SEM-ORG"),
                        row=(row.get("row") or ""),
                    )[:255]
                except Exception as e:
                    return self.jsonify_error('400', f"Invalid scan name template: {e}")
                result = self._start_scan_instance(scanname, ip_text, "", "", usecase)
                total_requested += 1
                row_count += 1

                if result[0] == "SUCCESS":
                    created.append({
                        "scan_id": result[1],
                        "target": ip_text,
                        "scan_name": scanname,
                        "network": network_text,
                    })
                else:
                    failures.append({
                        "network": network_text,
                        "target": ip_text,
                        "error": result[1],
                    })

            if max_total_scans and total_requested >= max_total_scans:
                break

        return {
            "status": "SUCCESS" if not failures else "PARTIAL",
            "created_count": len(created),
            "failed_count": len(failures),
            "created": created,
            "failures": failures,
        }

    @cherrypy.expose
    def startscan(
        self: 'SpiderFootWebUi',
        scanname: str,
        scantarget: str,
        modulelist: str = "",
        typelist: str = "",
        usecase: str = "",
        preset: str = None,
    ) -> str:
        """Initiate a scan.

        Args:
            scanname (str): scan name
            scantarget (str): scan target
            modulelist (str): comma separated list of modules to use
            typelist (str): selected modules based on produced event data types
            usecase (str): selected module group (passive, investigate, footprint, all)

        Returns:
            str: start scan status as JSON

        Raises:
            HTTPRedirect: redirect to new scan info page
        """
        result = self._start_scan_instance(scanname, scantarget, modulelist, typelist, usecase, preset=preset)
        if result[0] != "SUCCESS":
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(result).encode('utf-8')
            return self.error(result[1])

        if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
            cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
            return json.dumps(result).encode('utf-8')

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={result[1]}")

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def stopscan(self: 'SpiderFootWebUi', id: str) -> str:
        """Stop a scan.

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"Scan {scan_id} does not exist")

            scan_status = res[5]

            if scan_status == "FINISHED":
                return self.jsonify_error('400', f"Scan {scan_id} has already finished.")

            if scan_status == "ABORTED":
                return self.jsonify_error('400', f"Scan {scan_id} has already aborted.")

            if scan_status != "RUNNING" and scan_status != "STARTING":
                return self.jsonify_error('400', f"The running scan is currently in the state '{scan_status}', please try again later or restart SpiderFoot.")

        for scan_id in ids:
            dbh.scanInstanceSet(scan_id, status="ABORT-REQUESTED")

        return ""

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def vacuum(self):
        dbh = SpiderFootDb(self.config)
        try:
            if dbh.vacuumDB():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "Vacuuming the database failed"]).encode('utf-8')
        except Exception as e:
            return json.dumps(["ERROR", f"Vacuuming the database failed: {e}"]).encode('utf-8')

    #
    # DATA PROVIDERS
    #

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlog(self: 'SpiderFootWebUi', id: str, limit: str = None, rowId: str = None, reverse: str = None) -> list:
        """Scan log data.

        Args:
            id (str): scan ID
            limit (str): TBD
            rowId (str): TBD
            reverse (str): TBD

        Returns:
            list: scan log
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanLogs(id, limit, rowId, reverse)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], row[2], html.escape(row[3]), row[4]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanerrors(self: 'SpiderFootWebUi', id: str, limit: str = None) -> list:
        """Scan error data.

        Args:
            id (str): scan ID
            limit (str): limit number of results

        Returns:
            list: scan errors
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanErrors(id, limit)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], html.escape(str(row[2]))])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlist(self: 'SpiderFootWebUi') -> list:
        """Produce a list of scans.

        Returns:
            list: scan list
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceList()
        retdata = []

        for row in data:
            created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[3]))
            riskmatrix = {
                "HIGH": 0,
                "MEDIUM": 0,
                "LOW": 0,
                "INFO": 0
            }
            correlations = dbh.scanCorrelationSummary(row[0], by="risk")
            if correlations:
                for c in correlations:
                    riskmatrix[c[0]] = c[1]

            if row[4] == 0:
                started = "Not yet"
            else:
                started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[4]))

            if row[5] == 0:
                finished = "Not yet"
            else:
                finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[5]))

            retdata.append([row[0], row[1], row[2], created, started, finished, row[6], row[7], riskmatrix])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanstatus(self: 'SpiderFootWebUi', id: str) -> list:
        """Show basic information about a scan, including status and number of each event type.

        Args:
            id (str): scan ID

        Returns:
            list: scan status
        """
        dbh = SpiderFootDb(self.config)
        data = dbh.scanInstanceGet(id)

        if not data:
            return []

        created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[2]))
        started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[3]))
        ended = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[4]))
        riskmatrix = {
            "HIGH": 0,
            "MEDIUM": 0,
            "LOW": 0,
            "INFO": 0
        }
        correlations = dbh.scanCorrelationSummary(id, by="risk")
        if correlations:
            for c in correlations:
                riskmatrix[c[0]] = c[1]

        return [data[0], data[1], created, started, ended, data[5], riskmatrix]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanprogress(self: 'SpiderFootWebUi', id: str) -> dict:
        """Return execution progress summary for a scan."""
        if not id:
            return self.jsonify_error('404', "No scan specified")

        try:
            progress = self.scan_progress_summary(id)
        except Exception as e:
            return self.jsonify_error('500', str(e))

        if not progress:
            return self.jsonify_error('404', "Scan ID not found.")

        return progress

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scansummary(self: 'SpiderFootWebUi', id: str, by: str) -> list:
        """Summary of scan results.

        Args:
            id (str): scan ID
            by (str): filter by type

        Returns:
            list: scan summary
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        try:
            scandata = dbh.scanResultSummary(id, by)
        except Exception:
            return retdata

        try:
            statusdata = dbh.scanInstanceGet(id)
        except Exception:
            return retdata

        for row in scandata:
            if row[0] == "ROOT":
                continue
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[2]))
            retdata.append([row[0], row[1], lastseen, row[3], row[4], statusdata[5]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scancorrelations(self: 'SpiderFootWebUi', id: str) -> list:
        """Correlation results from a scan.

        Args:
            id (str): scan ID

        Returns:
            list: correlation result list
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        try:
            corrdata = dbh.scanCorrelationList(id)
        except Exception:
            return retdata

        for row in corrdata:
            retdata.append([row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresults(self: 'SpiderFootWebUi', id: str, eventType: str = None, filterfp: bool = False, correlationId: str = None) -> list:
        """Return all event results for a scan as JSON.

        Args:
            id (str): scan ID
            eventType (str): filter by event type
            filterfp (bool): remove false positives from search results
            correlationId (str): filter by events associated with a correlation

        Returns:
            list: scan results
        """
        retdata = []

        dbh = SpiderFootDb(self.config)

        if not eventType:
            eventType = 'ALL'

        try:
            data = dbh.scanResultEvent(id, eventType, filterfp, correlationId=correlationId)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            retdata.append([
                lastseen,
                html.escape(row[1]),
                html.escape(row[2]),
                row[3],
                row[5],
                row[6],
                row[7],
                row[8],
                row[13],
                row[14],
                row[4],
                row[10],
                row[11]
            ])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def findingdetail(self: 'SpiderFootWebUi', id: str, resulthash: str) -> dict:
        if not id or not resulthash:
            return self.jsonify_error('400', "Parâmetros obrigatórios ausentes.")

        dbh = SpiderFootDb(self.config)
        finding_data = None
        try:
            rows = dbh.scanElementSourcesDirect(id, [resulthash])
            if rows:
                row = rows[0]
                finding_data = {
                    "generated": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0])),
                    "data": row[1],
                    "source_data": row[2],
                    "module": row[3],
                    "event_type": row[4],
                    "confidence": row[5],
                    "visibility": row[6],
                    "risk": row[7],
                    "hash": row[8],
                    "source_hash": row[9],
                    "event_label": row[10],
                    "entity_type": row[11],
                    "false_positive": row[13]
                }
            state = dbh.findingStateGet(id, resulthash)
            evidence = dbh.findingEvidenceList(id, resulthash)
            validations = dbh.validationRunList(id, resulthash)
        except Exception as e:
            return self.jsonify_error('500', f"Erro ao carregar o achado: {e}")

        return {
            "finding": finding_data,
            "state": state,
            "validator_support": self.finding_validator_engine.describe_support(finding_data["event_type"] if finding_data else ""),
            "evidence": [
                {
                    "id": row[0],
                    "evidence_type": row[1],
                    "title": row[2],
                    "content": row[3],
                    "created": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[4]))
                } for row in evidence
            ],
            "validations": [
                {
                    "id": row[0],
                    "validator": row[1],
                    "status": row[2],
                    "summary": row[3],
                    "details": row[4],
                    "created": time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[5]))
                } for row in validations
            ],
            "labels": {
                "triage_status": self.finding_status_labels(),
                "relevance": self.finding_relevance_labels(),
                "exploitability": self.finding_exploitability_labels(),
                "analyst_verdict": self.analyst_verdict_labels()
            }
        }

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def findingupdate(self: 'SpiderFootWebUi', id: str, resulthash: str, triage_status: str, relevance: str,
                      exploitability: str, analyst_verdict: str, notes: str = "") -> list:
        if not id or not resulthash:
            return ["ERROR", "Parâmetros obrigatórios ausentes."]

        dbh = SpiderFootDb(self.config)
        try:
            dbh.findingStateSet(id, resulthash, triage_status, relevance, exploitability, analyst_verdict, notes)
        except Exception as e:
            return ["ERROR", f"Falha ao salvar o achado: {e}"]
        return ["SUCCESS", "Achado atualizado com sucesso."]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def findingevidenceadd(self: 'SpiderFootWebUi', id: str, resulthash: str, evidence_type: str, title: str, content: str) -> list:
        if not id or not resulthash or not title or not content:
            return ["ERROR", "Preencha tipo, título e conteúdo da evidência."]

        dbh = SpiderFootDb(self.config)
        try:
            dbh.findingEvidenceAdd(id, resulthash, evidence_type or "nota", title, content)
        except Exception as e:
            return ["ERROR", f"Falha ao registrar evidência: {e}"]
        return ["SUCCESS", "Evidência adicionada com sucesso."]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def findingvalidate(self: 'SpiderFootWebUi', id: str, resulthash: str) -> list:
        if not id or not resulthash:
            return ["ERROR", "Parâmetros obrigatórios ausentes."]

        dbh = SpiderFootDb(self.config)
        try:
            rows = dbh.scanElementSourcesDirect(id, [resulthash])
            if not rows:
                return ["ERROR", "Achado não encontrado."]
            row = rows[0]
            validation = self._run_finding_validation(row[4], row[1])
            dbh.validationRunAdd(id, resulthash, validation["validator"], validation["status"], validation["summary"], validation["details"])
        except Exception as e:
            return ["ERROR", f"Falha ao validar achado: {e}"]

        return ["SUCCESS", validation["summary"]]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanopssummary(self: 'SpiderFootWebUi', id: str) -> dict:
        if not id:
            return self.jsonify_error('404', "Nenhuma varredura especificada.")

        dbh = SpiderFootDb(self.config)
        try:
            summary = dbh.scanFindingOperationalSummary(id)
            verdict = dbh.caseVerdictGet(id)
        except Exception as e:
            return self.jsonify_error('500', f"Erro ao montar resumo operacional: {e}")

        return {
            "summary": summary,
            "verdict": verdict,
            "labels": {
                "triage_status": self.finding_status_labels(),
                "relevance": self.finding_relevance_labels(),
                "exploitability": self.finding_exploitability_labels(),
                "analyst_verdict": self.analyst_verdict_labels()
            }
        }

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanverdictupdate(self: 'SpiderFootWebUi', id: str, verdict: str, summary: str = "") -> list:
        if not id:
            return ["ERROR", "Nenhuma varredura especificada."]

        dbh = SpiderFootDb(self.config)
        try:
            dbh.caseVerdictSet(id, verdict, summary)
        except Exception as e:
            return ["ERROR", f"Falha ao salvar veredito: {e}"]
        return ["SUCCESS", "Veredito final atualizado com sucesso."]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresultsunique(self: 'SpiderFootWebUi', id: str, eventType: str, filterfp: bool = False) -> list:
        """Return unique event results for a scan as JSON.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            filterfp (bool): remove false positives from search results

        Returns:
            list: unique search results
        """
        dbh = SpiderFootDb(self.config)
        retdata = []

        try:
            data = dbh.scanResultEventUnique(id, eventType, filterfp)
        except Exception:
            return retdata

        for row in data:
            escaped = html.escape(row[0])
            retdata.append([escaped, row[1], row[2]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def search(self: 'SpiderFootWebUi', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search scans.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            value (str): filter search results by event value

        Returns:
            list: search results
        """
        try:
            return self.searchBase(id, eventType, value)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanhistory(self: 'SpiderFootWebUi', id: str) -> list:
        """Historical data for a scan.

        Args:
            id (str): scan ID

        Returns:
            list: scan history
        """
        if not id:
            return self.jsonify_error('404', "No scan specified")

        dbh = SpiderFootDb(self.config)

        try:
            return dbh.scanResultHistory(id)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanelementtypediscovery(self: 'SpiderFootWebUi', id: str, eventType: str) -> dict:
        """Scan element type discovery.

        Args:
            id (str): scan ID
            eventType (str): filter by event type

        Returns:
            dict
        """
        dbh = SpiderFootDb(self.config)
        pc = dict()
        datamap = dict()
        retdata = dict()

        # Get the events we will be tracing back from
        try:
            leafSet = dbh.scanResultEvent(id, eventType)
            [datamap, pc] = dbh.scanElementSourcesAll(id, leafSet)
        except Exception:
            return retdata

        # Delete the ROOT key as it adds no value from a viz perspective
        del pc['ROOT']
        retdata['tree'] = SpiderFootHelpers.dataParentChildToTree(pc)
        retdata['data'] = datamap

        return retdata
